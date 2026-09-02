"""Classify the invoice book into the deck's four revenue lines and build a
monthly EUR series. Amounts are 'Importo della fattura in EUR (Senza IVA)' -
net of the 22% Italian VAT, i.e. recognised revenue rather than cash collected."""
import openpyxl, datetime, collections, json, re

P="/root/.claude/uploads/d6aea6eb-35c4-5830-952d-9c99a50ab850/f0cb1d18-AI_Central____Anagrafiche_Fatture_3.xlsx"
wb=openpyxl.load_workbook(P, read_only=True, data_only=True)

SHEETS={ # name: (datecols, prov, eur_net, tipo, prod)
 "Entrate QONTO 2024-5": ([3],   4, 7, 10, 11),
 "Entrate QONTO 2026":   ([0,5], 3,10, 13, 14),
 "Entrate WISE":         ([1,6], 4,11, 14, 15),
}

def classify(prod, prov):
    s=(prod or "").lower(); v=(prov or "").lower()
    if "giroconto" in s or "giroconto" in v:            return "EXCLUDE_transfer"
    if "cashback" in s:                                 return "EXCLUDE_cashback"
    if "rimborso" in s or "rimborso" in v:              return "EXCLUDE_refund"
    if "ai library" in s or "subscription update" in s: return "Library"
    if ("network ads" in s or "boosts payout" in s
        or "refind ads" in s):                          return "Ad network"
    if "netline" in s or "tradepub" in s:               return "Lead-gen"
    # Predictable Source LLC is the Jobstream co-brand deal, not lead-gen
    # (confirmed by Alex, 27 Aug 2026)
    if "predictable source" in s or "jobstream" in s:   return "Sponsorship"
    if ("collaboration" in s or "quarter deal" in s or "dedicated issue" in s
        or "sponsorship" in s or "sponsored" in s or "linkedin" in s
        or "thought-leadership" in s or "partnership" in s):
                                                        return "Sponsorship"
    if "collabwork" in s or "affiliate" in s or "dub.co" in v: return "Affiliate"
    return "Other"

rows=[]
for name,(dcs,pc,ec,tc,prc) in SHEETS.items():
    for r in list(wb[name].iter_rows(values_only=True))[1:]:
        if not any(c is not None and str(c).strip()!="" for c in r): continue
        amt=r[ec] if ec<len(r) and isinstance(r[ec],(int,float)) else None
        if amt is None: continue
        d=next((r[i] for i in dcs if i<len(r) and isinstance(r[i],datetime.datetime)), None)
        prod=str(r[prc]).strip() if prc<len(r) and r[prc] else ""
        prov=str(r[pc]).strip()  if pc<len(r)  and r[pc]  else ""
        rows.append(dict(sheet=name, d=d, amt=float(amt), prod=prod, prov=prov,
                         line=classify(prod,prov)))

LINES=["Sponsorship","Ad network","Library","Lead-gen"]
excl   =[r for r in rows if r["line"].startswith("EXCLUDE")]
counted=[r for r in rows if not r["line"].startswith("EXCLUDE")]
dated  =[r for r in counted if r["d"]]
undated=[r for r in counted if not r["d"]]

print("RECONCILIATION (EUR, net of VAT)")
print(f"  all rows with an amount        {sum(r['amt'] for r in rows):>12,.2f}  n={len(rows)}")
for k,v in collections.Counter(r["line"] for r in excl).items():
    print(f"  less {k:26s} {sum(r['amt'] for r in excl if r['line']==k):>12,.2f}  n={v}")
print(f"  = revenue                      {sum(r['amt'] for r in counted):>12,.2f}  n={len(counted)}")
print(f"      of which dated             {sum(r['amt'] for r in dated):>12,.2f}  n={len(dated)}")
print(f"      of which UNDATED (QONTO)   {sum(r['amt'] for r in undated):>12,.2f}  n={len(undated)}")

print("\nBY LINE - all revenue")
for L in LINES+["Affiliate","Other"]:
    tot=sum(r["amt"] for r in counted if r["line"]==L)
    dat=sum(r["amt"] for r in dated   if r["line"]==L)
    if tot: print(f"  {L:14s} {tot:>12,.2f}   (dated {dat:>11,.2f}, undated {tot-dat:>10,.2f})")

months=collections.defaultdict(lambda: collections.defaultdict(float))
for r in dated: months[f"{r['d']:%Y-%m}"][r["line"]]+=r["amt"]
print("\nMONTHLY (dated only)")
print(f"  {'month':8s}" + "".join(f"{L:>13s}" for L in LINES) + f"{'other':>11s}{'TOTAL':>12s}")
for m in sorted(months):
    b=months[m]; other=sum(v for k,v in b.items() if k not in LINES)
    print(f"  {m:8s}" + "".join(f"{b.get(L,0):>13,.0f}" for L in LINES)
          + f"{other:>11,.0f}{sum(b.values()):>12,.0f}")

json.dump({m:dict(months[m]) for m in sorted(months)}, open("/home/claude/build/revenue-monthly.json","w"), indent=1)

und=collections.Counter()
for r in undated: und[r["line"]]+=r["amt"]
print("\nUNDATED QONTO 2024-5 residual by line:",
      {k:round(v,2) for k,v in und.most_common()})

ws=wb["Refunds & Disputes"]; rr=list(ws.iter_rows(values_only=True))
ref=sum(c for r in rr[1:] for c in [r[9] if len(r)>9 and isinstance(r[9],(int,float)) else 0])
print(f"\nRefunds & Disputes sheet total (EUR net): {ref:,.2f}  (n={len(rr)-1} rows, tracked separately)")
